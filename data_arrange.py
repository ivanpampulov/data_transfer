import openpyxl


def copy_and_fill_data():
    # Open the source and target workbooks
    source_workbook = openpyxl.load_workbook(input())
    target_workbook = openpyxl.load_workbook(input())

    # Set the source and target sheets
    source_sheet = source_workbook[input()]
    target_sheet = target_workbook[input()]
    users_email_address = input()

    # Find the last row in the source sheet
    last_row = source_sheet.max_row

    # Set the ranges for data in the source sheet
    source_data_A = source_sheet["A2:A" + str(last_row)]
    source_data_B = source_sheet["B2:B" + str(last_row)]
    source_data_D = source_sheet["D2:D" + str(last_row)]
    source_data_F = source_sheet["F2:F" + str(last_row)]
    source_data_AF = source_sheet["AF2:AF" + str(last_row)]
    source_data_L = source_sheet["L2:L" + str(last_row)]
    source_data_X = source_sheet["X2:X" + str(last_row)]
    source_data_Z = source_sheet["Z2:Z" + str(last_row)]
    source_data_AA = source_sheet["AA2:AA" + str(last_row)]
    source_data_Y = source_sheet["Y2:Y" + str(last_row)]
    source_data_W = source_sheet["W2:W" + str(last_row)]

    # Check valid data:
    for i in range(2, last_row + 1):

        #if there is no email address
        if source_sheet.cell(row=i, column=9, value=None):
            source_sheet.delete_rows(i, amount=1)

        #if lead was already exported from another Sales
        if users_email_address != source_sheet.cell(row=i, column=11):
            source_sheet.delete_rows(i, amount=1)

        #if email address is not valid
        if '@' not in str(source_sheet.cell(row=i, column=9)):
            source_sheet.delete_rows(i, amount=1)

        #if there is no family name
        if source_sheet.cell(row=i, column=2, value=None):
            source_sheet.cell(row=i, column=2, value=source_sheet.cell(row=i, column=1))

    # Copy data from source to target
    for i in range(2, last_row + 1):
        target_sheet.cell(row=i, column=1, value=source_data_A[i - 2][0].value)
        target_sheet.cell(row=i, column=2, value=source_data_B[i - 2][0].value)
        target_sheet.cell(row=i, column=3, value=source_data_D[i - 2][0].value)
        target_sheet.cell(row=i, column=9, value=source_data_F[i - 2][0].value)
        target_sheet.cell(row=i, column=7, value=source_data_AF[i - 2][0].value)
        target_sheet.cell(row=i, column=10, value=source_data_L[i - 2][0].value)
        target_sheet.cell(row=i, column=11, value=source_data_X[i - 2][0].value)
        target_sheet.cell(row=i, column=13, value=source_data_Z[i - 2][0].value)
        target_sheet.cell(row=i, column=17, value=source_data_AA[i - 2][0].value)
        target_sheet.cell(row=i, column=15, value=source_data_Y[i - 2][0].value)
        target_sheet.cell(row=i, column=18, value=source_data_W[i - 2][0].value)

    #Check for correct country
    for i in range(2, last_row + 1):
        if target_sheet.cell(row=i, column=7) == 'United States':
            target_sheet.cell(row=i, column=7, value='USA')


    # Fill other specified columns in the target sheet
    for i in range(2, last_row + 1):
        target_sheet.cell(row=i, column=4, value="Valya Krasteva")
        target_sheet.cell(row=i, column=5, value="Ivan Pampulov")
        target_sheet.cell(row=i, column=6,
                          value="Fun Walls, Ropes Course, Rollglider, Adventure Trail, Caving, Cloud Climb, Zip Line, Ninja Course, Tree Course, Slides")
        target_sheet.cell(row=i, column=8, value="Active Entertainment")
        target_sheet.cell(row=i, column=14, value="LinkedIn")
        target_sheet.cell(row=i, column=19, value="I first contacted client")
        target_sheet.cell(row=i, column=23, value="Global")

    # Save the changes to the target workbook
    target_workbook.save("D:\IVAN\Lead Import.xls")


if __name__ == "__main__":
    copy_and_fill_data()