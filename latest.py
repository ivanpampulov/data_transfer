import openpyxl
import openai

# Prompt for the source location and file name
sourceLocation = input("Enter source file location: ")
sourceFileName = input("Enter source file name: ")

# Open the source workbook
sourceWorkbook = openpyxl.load_workbook(f"{sourceLocation}/{sourceFileName}.xlsx")
sourceSheetName = input("Enter source sheet name: ")
sourceSheet = sourceWorkbook[sourceSheetName]

# Create a new target workbook
targetWorkbook = openpyxl.Workbook()
targetSheet = targetWorkbook.active

# Find the last row in the source sheet
lastRow = sourceSheet.max_row

# Set the ranges for data in the source sheet
sourceDataFirstName = sourceSheet['A2:A' + str(lastRow)]
sourceDataLastName = sourceSheet['B2:B' + str(lastRow)]
sourceDataCompanyName = sourceSheet['D2:D' + str(lastRow)]
sourceDataEmail = sourceSheet['F2:F' + str(lastRow)]
sourceDataCountry = sourceSheet['AF2:AF' + str(lastRow)]
sourceDataMobile = sourceSheet['L2:L' + str(lastRow)]
sourceDataDescription = sourceSheet['X2:X' + str(lastRow)]
sourceDataWebsite = sourceSheet['Z2:Z' + str(lastRow)]
sourceDataCLinkedIn = sourceSheet['AA2:AA' + str(lastRow)]
sourceDataPLinkedIn = sourceSheet['Y2:Y' + str(lastRow)]
sourceDataIndustry = sourceSheet['W2:W' + str(lastRow)]

# Create header in the target sheet
targetSheet.append([
    "First Name", "Last Name", "Account Name", "HQ Sales", "Assigned User Name",
    "Product Interests", "Country", "Division", "Email Address", "Mobile Phone",
    "Description", "Title", "Website", "Lead Source", "Person Linkedin Url",
    "WeChat", "Company Linkedin Url", "Industry", "Who Initiated the first contact?",
    "Department", "Reports To", "Facebook Account", "Teams"
])

# Copy data from source to target
for i in range(2, lastRow + 1):
    targetSheet.append([
        sourceDataFirstName[i - 2][0].value, sourceDataLastName[i - 2][0].value, sourceDataCompanyName[i - 2][0].value,
        "", "", "Fun Walls, Ropes Course, Rollglider, Adventure Trail, Caving, Cloud Climb, Zip Line, Ninja Course, Tree Course, Slides",
        sourceDataCountry[i - 2][0].value, "Active Entertainment", sourceDataEmail[i - 2][0].value,
        sourceDataMobile[i - 2][0].value, sourceDataDescription[i - 2][0].value, "", sourceDataWebsite[i - 2][0].value, "LinkedIn", sourceDataPLinkedIn[i - 2][0].value,
        "", sourceDataCLinkedIn[i - 2][0].value, sourceDataIndustry[i - 2][0].value, "I first contacted client", "", "", "",
        "Global"
    ])

HQSales = input("Who is the head sales: ")
assignedTo = input("To whom is this lead assigned to: ")

# Fill specified columns in the target sheet
for row in targetSheet.iter_rows(min_row=2, max_row=lastRow, min_col=4, max_col=5):
    for cell in row:
        cell.value = HQSales if cell.column == 4 else assignedTo

# Delete rows in the target sheet based on conditions
for i in range(lastRow, 1, -1):
    if targetSheet.cell(row=i, column=9).value is None or "@" not in str(targetSheet.cell(row=i, column=9).value):
        targetSheet.delete_rows(i, amount=1)
    if targetSheet.cell(row=i, column=2).value is None:
        targetSheet.cell(row=i, column=2, value=targetSheet.cell(row=i, column=1).value)

# Replace country names
country_replacements = {
    "United States": "USA",
    "South Korea": "Korea, South",
    "Republic of Indonesia": "Indonesia",
    "Holland": "Netherlands",
    "Czechia": "Czech Republic",
    "Macau": "Macao",
    "Antigua": "Barbuda",
    "The Bahamas": "Bahamas",
    "Republic of the Union of Myanmar": "Myanmar",
    "Suriname": "Surinam",
    "The Gambia": "Gambia",
    "Trinidad and Tobago": "Trinidad",
    "U.S. Virgin Islands": "US Virgin Islands",
    "New Caledonia": "New Caladonia",
    "Jersey": "United Kingdom",
    "Isle of man": "United Kingdom",
    "Guernsey": "United Kingdom",
    "Togo": "Ghana"
}

for i in range(2, lastRow + 1):
    if targetSheet.cell(row=i, column=7).value in country_replacements:
        targetSheet.cell(row=i, column=7).value = country_replacements[targetSheet.cell(row=i, column=7).value]

#Replace industry
industry_replacements = {
    "real estate": "Real Estate Developer",
    "architecture & planning": "Design",
    "civil engineering": "Engineering",
    "building materials": "Engineering",
    'glass, ceramics & concrete': "Manufacturing",
    "commercial real estate": 'Real Estate Developer',
    'textiles': "Manufacturing",
    'animation': 'Consulting',
    'farming': 'Manufacturing'
}

industry_list = ["Real Estate Developer",
                 "Design",
                 "Engineering",
                 "Family Entertainment Center",
                 "Seaside Resort",
                 "Airport",
                 "Family Resort",
                 "Consulting",
                 "Time shared resort"
                 ]

for i in range(2, lastRow + 1):
    if targetSheet.cell(row=i, column=18).value in industry_replacements:
        targetSheet.cell(row=i, column=18).value = industry_replacements[targetSheet.cell(row=i, column=18).value]

    else:
        openai.api_key = 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

        question = f'To which of the following industries is {targetSheet.cell(row=i, column=18).value} mostly related. Make a single choice. {", ".join(industry_list)}'
        response = openai.Completion.create(
            engine="davinci",
            prompt=question,
            max_tokens=1
        )
        answer = response.choices[0].text.strip()
        targetSheet.cell(row=i, column=18).value = answer
        industry_replacements[targetSheet.cell(row=i, column=18).value] = answer

# Prompt for the target location and file name
targetLocation = input("Where should I save the file: ")
targetFileName = input("Pick a name: ")

# Save the target workbook with the specified location and file name
targetWorkbook.save(f"{targetLocation}/{targetFileName}.xlsx")

# Close the workbooks
targetWorkbook.close()
sourceWorkbook.close()
